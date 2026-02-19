# ExcelExporter.py
import openpyxl
from openpyxl.styles import PatternFill
import os

BLOB_BASE_URL = "https://images202503.blob.core.windows.net/images/Pearson/Maths/"

# Columns A–Y matching the database template exactly
COLUMNS = [
    "exam",                  # A
    "year",                  # B
    "subject",               # C
    "level",                 # D
    "paper",                 # E
    "section",               # F
    "question_no",           # G
    "part",                  # H
    "sub_part",              # I
    "question_text",         # J
    "question_image",        # K  ← full question image (entire Q, repeated per part row)
    "question_total_marks",  # L
    "part_text",             # M
    "part_image",            # N  ← individual part image
    "marking_scheme_text",   # O
    "marking_scheme_image",  # P
    "part_total_marks",      # Q
    "low_credit_marks",      # R
    "low_credit_criteria",   # S
    "mid_credit_marks",      # T
    "mid_credit_criteria",   # U
    "high_credit_marks",     # V
    "high_credit_criteria",  # W
    "context_text",          # X
    "full_credit_1",         # Y
]

# Columns to highlight green (auto-filled by the pipeline)
GREEN_COLUMNS = {
    "exam", "year", "subject", "level", "paper",
    "question_no", "part", "sub_part",
    "question_image",   # K
    "part_image",       # N
}

GREEN_FILL = PatternFill(start_color="D9EAD3", end_color="D9EAD3", fill_type="solid")


class ExcelExporter:
    def __init__(self):
        self.rows = []

    def add_row(
        self,
        year, subject, paper, level,
        question_no, part, sub_part,
        part_image_filename,
        question_image_filename,   # full question image (may equal part_image when no parts)
    ):
        """
        Call this every time an image is saved in TaskPipeline.

        part_image_filename      – the cropped image for this specific part / sub-part
        question_image_filename  – the full question image (all parts combined);
                                repeated on every row that belongs to the same question
        """
        part_image_url     = BLOB_BASE_URL + part_image_filename
        question_image_url = BLOB_BASE_URL + question_image_filename

        row = {col: "" for col in COLUMNS}
        row["exam"]            = "UK A Level"
        row["year"]            = year
        row["subject"]         = subject
        row["level"]           = level
        row["paper"]           = paper
        row["question_no"]     = question_no
        row["part"]            = part if part else "NA"
        row["sub_part"]        = sub_part if sub_part else "NA"
        row["question_image"]  = question_image_url   # column K
        row["part_image"]      = part_image_url       # column N
        self.rows.append(row)

    def save(self, output_path):
        """Write all rows to Excel once processing is complete."""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Questions"

        # Header row
        for col_idx, col_name in enumerate(COLUMNS, start=1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            if col_name in GREEN_COLUMNS:
                cell.fill = GREEN_FILL

        # Data rows
        for row_idx, row_data in enumerate(self.rows, start=2):
            for col_idx, col_name in enumerate(COLUMNS, start=1):
                ws.cell(row=row_idx, column=col_idx, value=row_data.get(col_name, ""))

        # Auto column widths
        for col_idx, col_name in enumerate(COLUMNS, start=1):
            max_len = max(
                len(col_name),
                max((len(str(row.get(col_name, ""))) for row in self.rows), default=0),
            )
            ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = min(
                max_len + 2, 60
            )

        wb.save(output_path)
        print(f"\nExcel saved: {output_path}  ({len(self.rows)} rows)")
